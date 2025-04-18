# C:\Mail.ru\CodeIt\Django\BalkanPearl\BalkanPearlProject\BalkanPearlApp\admin.py
# -*- coding: utf-8 -*-
import datetime
from django.contrib import admin
from django.core.exceptions import ValidationError
from django.urls import path, reverse
from django.utils.encoding import force_str
from openpyxl.styles import Font, PatternFill
from decimal import Decimal

from BalkanPearlApp.models import Hotel, Address, HotelPhoto, WindowView, ApartmentType, Apartment, Season, \
    ApartmentPhoto, Booking, Review, BlogPost, SiteImage, Payment, Refund
from django import forms
from django.utils.translation import gettext_lazy as _
from django.shortcuts import render, HttpResponse
from django.db.models import Sum, Q
from django.http import HttpResponseForbidden
from openpyxl import Workbook
from .forms import AvailabilityReportFilterForm  # добавлено


# Register your models here.

@admin.register(Address)
class AddressAdmin(admin.ModelAdmin):
    list_display = ('street', 'city', 'state', 'postal_code', 'country')


class HotelAdminForm(forms.ModelForm):
    class Meta:
        model = Hotel
        fields = '__all__'


@admin.register(Hotel)
class HotelAdmin(admin.ModelAdmin):
    form = HotelAdminForm  # Используем пользовательскую форму
    list_display = ('name', 'description', 'floors', 'address', 'contact_email',
                    'contact_phone', 'instagram_link', 'facebook_link',)
    list_filter = ('name',)


@admin.register(HotelPhoto)
class HotelPhotoAdmin(admin.ModelAdmin):
    list_display = ('hotel', 'photo', 'description', 'uploaded_at')


@admin.register(WindowView)
class WindowViewAdmin(admin.ModelAdmin):
    list_display = ('view',)


@admin.register(ApartmentType)
class ApartmentTypeAdmin(admin.ModelAdmin):
    list_display = ('name',)


@admin.register(Apartment)
class ApartmentAdmin(admin.ModelAdmin):
    list_display = ( 'number', 'type', 'is_closed', 'capacity', 'base_price_per_night', 'window_view', 'floor',  'balcony', 'rooms_quantity',
                    'mini_kitchen', 'bathrooms_quantity', 'beds_quantity', 'sofa', 'washing_machine', 'desk',
                    'iron_and_board', 'child_bed', 'microwave', 'square', 'hotel',
                    )
    list_filter = ( 'number', 'window_view', 'floor', 'type', 'balcony', 'rooms_quantity',
                   'base_price_per_night', 'is_closed', 'hotel',)
    ordering = ('number', 'floor', 'hotel', )


@admin.register(Season)
class SeasonAdmin(admin.ModelAdmin):
    list_display = ('name', 'start_date', 'end_date', 'price_multiplier')


@admin.register(ApartmentPhoto)
class ApartmentPhotoAdmin(admin.ModelAdmin):
    list_display = ('get_apartment_number', 'photo', 'description', 'uploaded_at')

    def get_apartment_number(self, obj):
        return obj.apartment.number
    get_apartment_number.short_description = _('Apartment Number')


@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    list_display = ['id', 'booking', 'payment_value', 'payment_method', 'payment_date', 'note']
    list_filter = ['payment_method', 'payment_date']

    def save_model(self, request, obj, form, change):
        try:
            obj.save()
        except ValidationError as e:
            form.add_error(None, e.message)
            self.message_user(request, _("Error: %s") % e.message, level='error')


@admin.register(Refund)
class RefundAdmin(admin.ModelAdmin):
    list_display = ('id', 'booking', 'refund_value', 'refund_date', 'note')
    list_filter = ('refund_date', 'booking')
    search_fields = ('booking__id', 'refund_value')

    def save_model(self, request, obj, form, change):
        try:
            obj.save()
        except ValidationError as e:
            form.add_error(None, e.message)
            self.message_user(request, _("Error: %s") % e.message, level='error', )


class BookingAdminForm(forms.ModelForm):
    class Meta:
        model = Booking
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        people_quantity = cleaned_data.get('people_quantity')
        apartments = cleaned_data.get('apartments')
        check_in = cleaned_data.get('check_in')
        check_out = cleaned_data.get('check_out')

        # Если какие-то поля отсутствуют, пропускаем проверку
        if not (apartments and people_quantity and check_in and check_out):
            return cleaned_data
        total_capacity = 0
        apartments_numbers = []
        for apt in apartments:
            # Подсчитываем суммарную вместимость номеров
            total_capacity += apt.capacity
            apartments_numbers.append(apt.number)
            # Проверка на перекрытие дат бронирований
            qs = Booking.objects.filter(apartments=apt, status='confirmed')
            if self.instance.pk:
                qs = qs.exclude(pk=self.instance.pk)
            qs = qs.filter(check_in__lt=check_out, check_out__gt=check_in)
            if qs.exists():
                raise ValidationError(f"Апартамент {apt.number} уже забронирован на выбранные даты")
        # Проверка вместимости
        if people_quantity > total_capacity:
            raise ValidationError(f"Количество гостей превышает вместимость апартаментов {apartments_numbers}")
        return cleaned_data


@admin.register(Booking)
class BookingAdmin(admin.ModelAdmin):
    form = BookingAdminForm
    list_display = ('id', 'get_apartments', 'user', 'check_in', 'check_out', 'status', 'total_value',
                    'debt', 'people_quantity', 'created_at', )
    list_filter = ('status', 'user')
    search_fields = ('user__username', 'status')
    readonly_fields = ('created_at', 'total_value', 'debt')
    actions = ['cancel_booking']


    def get_urls(self):
        urls = [
            path(
                'reports/',
                self.admin_site.admin_view(self.reports_view),
                name='BalkanPearlApp_booking_reports'  # Уточнённое имя URL
            ),
            path(
                'export-reports/',
                self.admin_site.admin_view(self.export_reports),
                name='BalkanPearlApp_export_reports'
            ),
            path(
                'availability-report/',
                self.admin_site.admin_view(self.availability_report_view),
                name='BalkanPearlApp_booking_availability_report'
            ),
        ]
        return urls + super().get_urls()  # Кастомные URL должны быть ПЕРЕД стандартными

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['report_url'] = reverse('admin:BalkanPearlApp_booking_reports')
        extra_context['availability_url'] = reverse('admin:BalkanPearlApp_booking_availability_report')
        return super().changelist_view(request, extra_context=extra_context)

    def cancel_booking(self, request, queryset):
        """Эта функция — кастомное действие (admin action) в Django Admin.
        Она позволяет администратору отменять выбранные бронирования через панель
        управления."""

        """self — текущий экземпляр класса BookingAdmin (наследник admin.ModelAdmin).
            request — объект запроса HTTP (от Django Admin).
            queryset — набор объектов Booking, выбранных администратором."""
        cancelled = queryset.exclude(status='cancelled')

        for booking in cancelled:
            booking.cancel_booking(cancelled_by='admin')
        self.message_user(request, _("Выбранные бронирования отменены."))

    cancel_booking.short_description = _("Отменить выбранные бронирования")

    def get_apartments(self, obj):
        if not obj.pk:
            return ""
        return ", ".join(str(apartment.number) for apartment in obj.apartments.all())
    get_apartments.short_description = _('Апартаменты')

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        booking = form.instance
        if booking.status == 'confirmed':
            if booking.manual_value is not None:
                booking.total_value = booking.manual_value
            else:
                booking.total_value = booking.calculate_total_value()
        else:
            booking.total_value = Decimal('0.00')
        booking.debt = booking.total_value
        booking.save(update_fields=['total_value', 'debt'])


    def availability_report_view(self, request):
        """
        Кастомное представление, которое:
        - если GET-параметр action=export, генерирует и возвращает файл Excel,
        - иначе отрисовывает HTML-отчет с формой фильтров и таблицей.
        """
        if not request.user.is_staff:
            return HttpResponseForbidden(_("Доступ запрещён"))

        action = request.GET.get('action', 'filter')  # либо "filter", либо "export"
        form = AvailabilityReportFilterForm(request.GET or None)
        context = {}

        # Получаем список апартаментов из GET-параметров
        selected_apartments_list = request.GET.getlist('apartments')
        context['selected_apartments'] = selected_apartments_list  # добавляем в контекст

        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            hotel = form.cleaned_data.get('hotel')
            selected_apartments = form.cleaned_data.get('apartments')

            # Формируем список дат (включительно)
            delta = (end_date - start_date).days + 1
            date_range = [start_date + datetime.timedelta(days=i) for i in range(delta)]

            # Фильтруем апартаменты
            apartments_qs = Apartment.objects.all()
            if hotel:
                apartments_qs = apartments_qs.filter(hotel=hotel)
            if selected_apartments:
                apartments_qs = apartments_qs.filter(pk__in=[apt.pk for apt in selected_apartments])

            # Формируем матрицу доступности:
            # Для каждого апартамента и для каждого дня определяем статус:
            #   - "booked"  – если есть бронирование со статусом confirmed,
            #   - "free"    – если нет бронирования,
            #   - "closed"  – если у апартамента is_closed=True.
            occupancy_matrix = []
            for apt in apartments_qs:
                row = {
                    'apartment': str(apt),
                    'occupancy': []
                }
                for day in date_range:
                    if apt.is_closed:
                        status = "closed"
                    else:
                        # Поскольку проверка происходит для конкретного дня,
                        # интервал для проверки – [day, day + 1)
                        day_end = day + datetime.timedelta(days=1)
                        booked = apt.bookings.filter(
                            check_in__lt=day_end,
                            check_out__gt=day,
                            status='confirmed'
                        ).exists()
                        status = "booked" if booked else "free"
                    row['occupancy'].append(status)
                occupancy_matrix.append(row)

            context.update({
                'form': form,
                'date_range': date_range,
                'occupancy_matrix': occupancy_matrix,
                'hotel_selected': hotel,
                'start_date': start_date,
                'end_date': end_date,
            })

            if action == 'export':
                return self.export_availability_report_excel(date_range, occupancy_matrix, hotel, start_date, end_date)

        else:
            context['form'] = form

        return render(request, 'admin/BalkanPearlApp/booking/availability_report.html', context)

    def export_availability_report_excel(self, date_range, occupancy_matrix, hotel, start_date, end_date):
        """
        Генерирует Excel-файл, где в ячейках:
        - 1 если апартамент забронирован,
        - 0 если свободен,
        - -1 если закрыт.
        Применяются заливки для наглядности.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = force_str(_("Отчет по доступности"))

        # Шапка: название отеля (если выбран) и период
        header_title = _("Отель: %(hotel)s, Период: %(start_date)s - %(end_date)s") % {
            "hotel": hotel.name if hotel else _("Все отели"),
            "start_date": start_date,
            "end_date": end_date,
        }
        header_title = force_str(header_title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(date_range) + 1)
        ws.cell(row=1, column=1, value=header_title)

        # Заголовки столбцов: апартаменты и даты
        ws.cell(row=2, column=1, value=force_str(_("Апартамент")))
        for col_index, day in enumerate(date_range, start=2):
            ws.cell(row=2, column=col_index, value=day.strftime('%Y-%m-%d'))

        # Определяем стили заливки для ячеек
        fill_booked = PatternFill(fill_type="solid", start_color="FFCCCC", end_color="FFCCCC")
        fill_free = PatternFill(fill_type="solid", start_color="CCFFCC", end_color="CCFFCC")
        fill_closed = PatternFill(fill_type="solid", start_color="D3D3D3", end_color="D3D3D3")

        # Заполнение данными: для каждого апартамента по каждой дате
        row_number = 3
        for row in occupancy_matrix:
            ws.cell(row=row_number, column=1, value=row['apartment'])
            for col_index, status in enumerate(row['occupancy'], start=2):
                cell = ws.cell(row=row_number, column=col_index)
                if status == "booked":
                    cell.value = 1
                    cell.fill = fill_booked
                elif status == "free":
                    cell.value = 0
                    cell.fill = fill_free
                elif status == "closed":
                    cell.value = -1
                    cell.fill = fill_closed
            row_number += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="availability_report.xlsx"'
        wb.save(response)
        return response

    def reports_view(self, request):
        if not request.user.is_staff:
            return HttpResponseForbidden(_('Доступ запрещен'))
        context = self._process_report(request)
        return render(request, 'admin/BalkanPearlApp/booking/reports.html', context)

    def export_reports(self, request):
        if not request.user.is_staff:
            return HttpResponseForbidden(_('Доступ запрещен'))
        context = self._process_report(request)
        return self._generate_excel(context)


    def _process_report(self, request):
        form = ReportFilterForm(request.GET or None)
        results = []
        total_income = 0
        total_debt = 0
        planned_income = 0

        if form.is_valid():
            filters = Q()
            hotel = form.cleaned_data['hotel']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            apartment = form.cleaned_data['apartment']
            status = form.cleaned_data['status']
            season = form.cleaned_data['season']

            bookings = Booking.objects.filter(
                Q(check_in__lte=end_date) &
                Q(check_out__gte=start_date)
            )
            if hotel:
                bookings = bookings.filter(apartments__hotel=hotel)
            if apartment:
                bookings = bookings.filter(apartments=apartment)
            if status:
                bookings = bookings.filter(status=status)
            if season:
                bookings = bookings.filter(
                    Q(check_in__lte=season.end_date) &
                    Q(check_out__gte=season.start_date)
                )

            # Расчёт фактического дохода
            payments = Payment.objects.filter(booking__in=bookings)
            refunds = Refund.objects.filter(booking__in=bookings)

            total_payments = payments.aggregate(total=Sum('payment_value'))['total'] or 0
            total_refunds = refunds.aggregate(total=Sum('refund_value'))['total'] or 0
            total_income = total_payments - total_refunds

            # Планируемый доход (по бронированиям)
            planned_income = bookings.filter(status='confirmed').aggregate(
                total=Sum('total_value')
            )['total'] or 0

            # Дебиторская задолженность
            total_debt = bookings.exclude(
                status__in=['cancelled_by_user', 'cancelled_by_admin']
            ).aggregate(total=Sum('debt'))['total'] or 0

            # Детализация по апартаментам
            for apt in Apartment.objects.all():
                apt_bookings = bookings.filter(apartments=apt)

                # Фактический доход
                apt_payments = Payment.objects.filter(
                    booking__in=apt_bookings
                ).aggregate(total=Sum('payment_value'))['total'] or 0
                apt_refunds = Refund.objects.filter(
                    booking__in=apt_bookings
                ).aggregate(total=Sum('refund_value'))['total'] or 0

                # Планируемый доход
                apt_planned = apt_bookings.filter(status='confirmed').aggregate(
                    total=Sum('total_value')
                )['total'] or 0

                apt_debt = apt_bookings.exclude(
                    status__in=['cancelled_by_user', 'cancelled_by_admin']
                ).aggregate(total=Sum('debt'))['total'] or 0

                if apt_payments or apt_debt:
                    results.append({
                        'apartment': str(apt.number),  # Явное преобразование в строку
                        'actual_income': float(apt_payments - apt_refunds),
                        'planned_income': float(apt_planned),
                        'debt': float(apt_debt)
                    })

        return {
            'form': form,
            'results': results,
            'total_income': total_income,
            'planned_income': planned_income,
            'total_debt': total_debt,
            'params': request.GET.urlencode()
        }

    def _generate_excel(self, context):
        wb = Workbook()
        ws = wb.active
        ws.title = force_str(_("Financial Report"))  # Явное преобразование

        # Заголовки
        headers = [
            u"Апартамент",
            u"Фактический доход",
            u"Планируемый доход",
            u"Дебиторская задолженность"
        ]
        ws.append(headers)

        # Форматирование заголовков
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font

        # Данные
        for item in context['results']:
            row = [
                force_str(item['apartment']),
                float(item['actual_income']),
                float(item['planned_income']),
                float(item['debt'])
            ]
            ws.append(row)

        # Автонастройка ширины столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            charset='utf-8'
        )
        response['Content-Disposition'] = 'attachment; filename="finance_report.xlsx"'
        wb.save(response)
        return response

    @admin.display(description=_('Задолженность'))
    def debt_display(self, obj):
        return obj.debt


    def get_fields(self, request, obj=None, **kwargs):
        # Не исключаем поле apartments, чтобы его можно было выбрать при создании бронирования
        return super().get_fields(request, obj, **kwargs)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if obj is None and 'apartments' in form.base_fields:  # добавление нового объекта
            # Если объект еще не сохранён, используем стандартный SelectMultiple
            form.base_fields['apartments'].widget = forms.SelectMultiple()
        return form




@admin.register(Review)
class ReviewAdmin(admin.ModelAdmin):
    list_display = ('user', 'apartment', 'rating', 'commentary', 'created_at', 'anonymous')


@admin.register(BlogPost)
class BlogPostAdmin(admin.ModelAdmin):
    list_display = ('title', 'content', 'author', 'created_at',)


@admin.register(SiteImage)
class SiteImageAdmin(admin.ModelAdmin):
    list_display = ('image', 'description')


class ReportFilterForm(forms.Form):
    start_date = forms.DateField(
        label=_('Начало периода'),
        widget=forms.DateInput(attrs={'type': 'date'}),
        initial=datetime.date.today().replace(day=1)
    )
    end_date = forms.DateField(
        label=_('Конец периода'),
        widget=forms.DateInput(attrs={'type': 'date'}),
        initial=datetime.date.today()
    )
    hotel = forms.ModelChoiceField(
        queryset=Hotel.objects.all(),
        required=False,
        label=_('Отель'),
    )
    apartment = forms.ModelChoiceField(
        queryset=Apartment.objects.all(),
        required=False,
        label=_('Апартамент'),
    )
    status = forms.ChoiceField(
        choices=Booking.STATUS_CHOICES,
        required=False,
        label=_('Статус бронирования')
    )
    season = forms.ModelChoiceField(
        queryset=Season.objects.all(),
        required=False,
        label=_('Сезон')
    )
